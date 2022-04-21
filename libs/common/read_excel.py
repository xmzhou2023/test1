'''
进行excel数据驱动,excel数据驱动类底层代码
'''
import logging

from openpyxl import Workbook,load_workbook
from libs.config.conf import PEROJECT_PATH
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font,colors,Alignment
import openpyxl
from libs.config.conf import TESTCASE_PATH
from libs.common.logger_ui import log
import csv

def excel(path):
    """读取excel"""
    excel = load_workbook(path)
    # 获取sheet页
    sheets = excel.sheetnames
    # 遍历sheet页
    for sheet in sheets:
        # 判断该表格是否为测试用例：根据表格sheet名是否含有testcase
        if 'TestCase' in sheet:
            sheet_temp = excel[sheet]

            for i, d in enumerate(sheet_temp.iter_rows(values_only=True)):
                if i >= 4:
                    _data = dict(步骤名=d[4],关键字=d[5],参数=[_ for _ in d[6:9] if _])
                    yield _data
        else:
            log.info('[ {} ]表格内容不是测试用例'.format(sheet))
            # 写入后进行保存
if __name__ == '__main__':
    excel(TESTCASE_PATH)
