'''
进行excel数据驱动,excel数据驱动类底层代码
'''
import openpyxl
from project.IPM.page_base.pathconfig import *
pa=os.path.join(DATA, 'ProcessCenter_MaterialRequisition_Add.xlsx')


def read_excel():
    workbook = openpyxl.load_workbook(pa)
    sheet=workbook['Sheet']
    minrow=sheet.min_row
    maxrow=sheet.max_row
    mincol=sheet.min_column
    maxcol=sheet.max_column
    for m in range(mincol, maxcol + 1):
        for n in range(minrow, maxrow + 1):
            cell = sheet.cell(n, m).value
            print(cell,end=" ")

        print()

def read_data(filename,sheetname):
    #打开工作簿
    workbook = openpyxl.load_workbook(filename)
    sh=workbook[sheetname]
    #取到每行数据放入元组中
    datas=list(sh.rows)
    #拿到表头第一行数据
    title=[i.value for i in datas[0]]
    #定义一个空的列表用来接收所有的用例
    cases=[]
    #取除了第一行以外的所有的数据
    def validate(cell):
        return None if cell.value == None else cell.value
    for i in datas[1:]:
        values = map(validate, i)
        #把title表头变为键、把values变为值，然后通过zip函数进行配对，并且打包放入字典中

        case=dict(zip(title,values))
        cases.append(case)
    #运行用例
    return cases
if __name__ == '__main__':
    res=read_data(pa,'Sheet')
    print(res)
    for id in res:
        print(id.get("测试"))
