import logging
from openpyxl import load_workbook
from libs.common.read_element import Element
from public.base.basics import Base, random_list
from libs.common.time_ui import sleep
from libs.config.conf import BASE_DIR
from ..test_case.conftest import *
import random

object_name = os.path.basename(__file__).split('.')[0]
user = Element(pro_name, object_name)

class UserManagementPage(Base):
    """UserManagementPage 页面元素类"""

    @allure.step("user management页面，点击Add新增按钮")
    def click_add_user(self):
        self.is_click(user['Add'])
        sleep(2.5)

    @allure.step("进入Add user页面， 选择传音员工类型")
    def click_staff_type_value(self, type1):
        self.presence_sleep_dcr(user['Staff Type'])
        self.is_click(user['Staff Type'])
        sleep(1)
        self.is_click(user['Transsion Staff'], type1)

    @allure.step("Add user页面， 输入user ID字段")
    def input_user_id(self, content):
        self.is_click(user['User ID'])
        self.input_text_dcr(user['User ID'], txt=content)
        sleep(1.5)

    @allure.step("Add user页面， 输入user Name字段")
    def input_add_user_name(self, content):
        self.presence_sleep_dcr(user['Add User Name'])
        self.is_click(user['Add User Name'])
        self.input_text(user['Add User Name'], txt=content)

    @allure.step("Edit user页面， 输入user Name字段")
    def input_edit_user_name(self, content):
        self.presence_sleep_dcr(user['Edit User Name'])
        self.is_click(user['Edit User Name'])
        self.input_text(user['Edit User Name'], txt=content)


    @allure.step("Add user页面，点击user Name属性，释放光标")
    def click_user_name(self):
        self.is_click(user['Edit User Name'])

    @allure.step("Add user页面， 输入销售区域，然后选中输入的销售区域")
    def input_sales_region(self, content):
        self.is_click(user['Sales Region'])
        self.input_text(user['Sales Region'], txt=content)
        sleep(3)
        self.is_click(user['Sales Region Value'])

    @allure.step("Add user页面， 输入国家城市，然后选中输入的国家城市")
    def input_country_city(self, content):
        self.is_click(user['Country City'])
        self.input_text(user['Country City'], txt=content)
        sleep(2)
        self.is_click(user['Country City Value'])

    @allure.step("Add user页面，选择点击品牌")
    def click_select_brand(self):
        self.is_click(user['Add Select Brand'])
        sleep(1)
        self.presence_sleep_dcr(user['Select Brand Value'], 'Infinix')
        self.is_click(user['Select Brand Value'], 'Infinix')
        self.is_click(user['Select Brand Value'], 'TECNO')
        self.is_click(user['Select Brand Value'], 'itel')

    @allure.step("关闭User Management菜单")
    def click_close_user_mgt(self):
        self.is_click(user['关闭用户管理菜单'])

    @allure.step("Edit user页面，选择点击品牌")
    def click_edit_trans_brand(self, edit_brand):
        self.is_click(user['Edit Select Brand'])
        sleep(1)
        self.presence_sleep_dcr(user['Select Brand Value'], edit_brand)
        self.is_click(user['Select Brand Value'], edit_brand)

    @allure.step("Edit user页面，选择点击品牌")
    def click_edit_dealer_brand(self):
        self.is_click(user['Edit Select Brand'])
        sleep(1)
        self.presence_sleep_dcr(user['Select Brand Value'], 'Infinix')
        self.is_click(user['Select Brand Value'], 'Infinix')

    @allure.step("Add user页面，输入职位，选中输入的职位")
    def input_position_transsion(self, content):
        self.is_click(user['Position'])
        self.input_text(user['Position'], content)
        sleep(1)
        self.presence_sleep_dcr(user['Position Value Transsion'], content)
        self.is_click(user['Position Value Transsion'], content)

    @allure.step("Add user页面，输入上级领导，选中输入的上级领导")
    def input_superior(self, content):
        self.is_click(user['Superior'])
        self.input_text(user['Superior'], content)
        sleep(2.5)
        self.presence_sleep_dcr(user['Superior Value'], "lhmadmin lhmadmin")
        self.is_click(user['Superior Value'], "lhmadmin lhmadmin")

    @allure.step("Add user页面，输入邮箱")
    def input_email(self, content):
        self.is_click(user['Email'])
        self.input_text(user['Email'], content)

    @allure.step("Add user页面，输入联系电话")
    def input_contact_no(self, content):
        self.input_text(user['Contact No'], content)

    @allure.step("Add user页面，选择性别")
    def click_gender_female(self, context):
        self.is_click(user['Gender'])
        sleep(0.5)
        self.presence_sleep_dcr(user['Gender Female'], context)
        self.is_click(user['Gender Female'], context)

    @allure.step("Add user页面，点击Submit提交按钮")
    def click_add_user_submit(self):
        self.is_click(user['Add User Submit'])
        sleep(0.8)

    @allure.step("获取列表User ID文本内容")
    def get_text_user_id(self):
        sleep(1)
        userid = self.element_text(user['获取列表文本User ID'])
        return userid

    @allure.step("获取User Management列表,No Data文本内容")
    def get_user_management_no_data(self):
        get_no_data = self.element_text(user['获取所有文本'])
        return get_no_data

    @allure.step("'获取列表User Name文本内容")
    def get_text_user_name(self):
        username = self.element_text(user['获取列表文本User Name'])
        return username

    @allure.step("点击搜索功能")
    def click_search(self):
        self.is_click(user['Search'])
        sleep(3.5)

    @allure.step("点击重置按钮")
    def click_reset(self):
        self.is_click(user['Reset'])
        sleep(3)

    @allure.step("新增Add 传音员工操作步骤")
    def add_trans_user_operation(self, staff_type, trans_userid, trans_username, sales_region, country_city, superior, position, email, contact_no, gender):
        self.click_add_user()
        self.click_staff_type_value(staff_type)
        self.input_user_id(trans_userid)
        self.input_add_user_name(trans_username)
        self.input_sales_region(sales_region)
        self.input_country_city(country_city)
        """选择3个品牌 """
        self.click_select_brand()
        self.input_superior(superior)
        self.input_position_transsion(position)
        self.input_email(email)
        self.input_contact_no(contact_no)
        self.click_gender_female(gender)
        self.click_add_user_submit()

    @allure.step("编辑Edit 传音员工基本信息操作步骤")
    def edit_trans_user_info_operation(self, edit_username, brand, email, contact_no, gender):
        self.click_edit()
        self.input_edit_user_name(edit_username)
        self.click_edit_trans_brand(brand)
        self.click_user_name()
        self.input_email(email)
        self.input_contact_no(contact_no)
        self.click_gender_female(gender)
        self.click_add_user_submit()

    @allure.step("新增Add 代理员工操作步骤")
    def add_dealer_user_operation(self, staff_type, belong_to_cust, user_name, sales_region, country_city, position, superior, email, contact_no, gender):
        self.click_add_user()
        self.click_staff_type_value(staff_type)
        self.input_belong_to_cust(belong_to_cust)
        self.input_add_user_name(user_name)
        self.input_sales_region(sales_region)
        self.input_country_city(country_city)
        """选择3个品牌"""
        self.click_select_brand()
        self.input_position_dealer(position)
        self.input_superior(superior)
        self.input_email(email)
        self.input_contact_no(contact_no)
        self.click_gender_female(gender)
        self.click_add_user_submit()


    """编辑用户时，筛选用户"""
    @allure.step("用户管理列表页面，筛选User，进行编辑或者删除")
    def input_query_User(self, userid):
        self.is_click_dcr(user['点击筛选用户输入框'])
        self.input_text_dcr(user['输入筛选用户输入框'], userid)

    @allure.step("点击编辑功能")
    def click_edit(self):
        self.presence_sleep_dcr(user['修改第一个Edit'])
        self.is_click_dcr(user['修改第一个Edit'])
        sleep(2)

    @allure.step("点击第一个checkbox,对用户进行辞职操作")
    def click_first_checkbox(self):
        self.is_click_dcr(user['勾选第一个复选框'])

    @allure.step("编辑用户提交成功提示语")
    def get_set_up_successfully(self):
        set_up_success = self.element_text(user['Set Up Successfully'])
        return set_up_success

    @allure.step("点击更多操作,点击离职功能")
    def click_more_option_quit(self):
        self.mouse_hover_click(user['More Option'])
        #self.is_click(user['More Option'])
        #sleep(2)
        self.presence_sleep_dcr(user['Quit'])
        self.is_click(user['Quit'])
        sleep(1.5)
        self.is_click_dcr(user['确认删除Yes'])


    @allure.step("点击更多操作,点击离职,弹出User Disable Setting 窗口，获取User名称断言")
    def get_user_disable_setting_username(self):
        get_user_disable = self.element_text(user['Get User Disable Setting User'])
        return get_user_disable

    @allure.step("获取删除成功提示语文本内容")
    def get_text_delete_success(self):
        del_success = self.element_text(user['Disabled Successfully'])
        return del_success


    @allure.step("进入Add user页面， 输入客户ID")
    def input_belong_to_cust(self, content):
        self.is_click(user['Belong To Customer'])
        self.input_text(user['Belong To Customer'], txt=content)
        sleep(2)
        self.is_click(user['Belong To Customer Value'], "UG4019912")

    @allure.step("随机生成userid")
    def user_id_random(self):
        num = str(random.randint(100, 999))
        userid = '202211' + num
        return userid

    @allure.step("随机生成username")
    def trans_user_name_random(self):
        num = str(random.randint(100, 999))
        username = "test_trans_user" + num
        return username

    @allure.step("随机生成username")
    def dealer_user_name_random(self):
        num = str(random.randint(100, 999))
        username = "test_dealer_user" + num
        return username

    @allure.step("随机生成电话号码尾号")
    def number_radom(self):
        num = str(random.randint(100, 999))
        return num

    @allure.step("Add user页面，输入职位，选中输入的职位")
    def input_position_dealer(self, content):
        self.is_click(user['Position'])
        self.input_text(user['Position'], txt=content)
        sleep(1)
        self.is_click(user['Position Value Dealer'], content)


    """查询列表用户"""
    @allure.step("用户管理页面，获取列表文本内容方法")
    def get_user_management_list_data(self, data):
        self.presence_sleep_dcr(user[data])
        get_data = self.element_text(user[data])
        return get_data

    @allure.step("用户管理页面，获取列表Total分页总数")
    def get_total(self):
        get_total = self.element_text(user['Get Total'])
        get_total1 = get_total[6:]
        return get_total1

    @allure.step("断言分页总数是否存在数据")
    def assert_total(self, total):
        if int(total) > 1000:
            logging.info("查看用户管理列表，分页总条数大于1000，用户总记录Total：{}".format(total))
        else:
            logging.info("查看用户管理列表，分页总条数小于1000，用户总记录Total：{}".format(total))

    @allure.step("断言 精确查询结果 User Management列表，字段列、字段内容是否与预期的字段内容值一致，有滚动条")
    def assert_user_management_field(self, header, content):
        DomAssert(self.driver).assert_search_result(user['表格字段'], user['表格指定列内容2'], header, content, sc_element=user['水平滚动条'])

    @allure.step("断言 模糊查询结果User Management列表，字段列、字段内容是否与预期的字段内容值一致，有滚动条")
    def assert_contains_user_management_field(self, header, content):
        DomAssert(self.driver).assert_search_contains_result(user['表格字段'], user['表格指定列内容2'], header, content, sc_element=user['水平滚动条'])

    @allure.step("Sql语句删除新增或导入的用户")
    def sql_delete_user(self, user_id):
        sql1 = SQL('DCR', 'test')
        sql1.delete_db(
            f"delete from t_user where USER_CODE = '{user_id}'")
        sql1.delete_db(
            f"delete from t_employee where EMP_CODE = '{user_id}'")

    """用户重置密码"""
    @allure.step("点击重置密码及重置密码确认功能")
    def click_more_reset_password(self):
        self.is_click(user['More Option'])
        sleep(2)
        self.presence_sleep_dcr(user['Reset Password'])
        self.is_click(user['Reset Password'])
        sleep(1.4)
        self.presence_sleep_dcr(user['Reset Password Yes'])
        self.is_click(user['Reset Password Yes'])

    @allure.step("登录时，弹出设置新密码窗口，获取New Password 标签")
    def get_new_password_label(self):
        get_new_password = self.element_text(user['Get New Password'])
        return get_new_password


    @allure.step("重置密码成功后，该用户登录时，弹出设置新密码窗口，输入新密码与确认新密码，并点击保存按钮")
    def input_new_password_save(self, new_password):
        self.is_click(user['Input New Password'])
        self.input_text(user['Input New Password'], txt=new_password)
        sleep(1)
        self.is_click(user['Input Confirm Password'])
        self.input_text(user['Input Confirm Password'], txt=new_password)
        sleep(0.5)
        self.is_click(user['New Password Save'])

    @allure.step("输入新密码登录，点击OK")
    def click_save_successfully_ok(self):
        self.is_click(user['Save successfully OK'])
        sleep(2)

    @allure.step("登录页面，输入新设置的密码")
    def input_login_password(self, password):
        self.is_click(user['login Input Password'])
        self.input_text(user['login Input Password'], txt=password)

    @allure.step("登录页面，点击Login登录按钮")
    def click_login(self):
        self.is_click(user['Click Login'])
        sleep(3.5)


    """导入用户操作"""
    @allure.step("User Management页面，点击Import 按钮")
    def click_import(self):
        self.is_click(user['Import Button'])
        sleep(1.5)

    @allure.step("User Management页面，点击Import Save 按钮")
    def click_import_save(self):
        self.is_click(user['Import Save'])

    @allure.step("User Management页面，点击Import 导入功能")
    def click_import_upload_save(self, file1):
        self.is_click(user['Add Upload'])
        sleep(4)
        ele = self.driver.find_element('xpath', "//button//..//input[@name='file']")
        ele.send_keys(file1)
        sleep(3)
        self.is_click(user['Import Save'])
        sleep(2)
        self.presence_sleep_dcr(user['Upload Confirm'])
        self.is_click(user['Upload Confirm'])


    @allure.step("User Management页面，获取Total分页总条数")
    def get_list_total(self):
        self.presence_sleep_dcr(user['Get Total'])
        get_total = self.element_text(user['Get Total'])
        get_total1 = get_total[6:]
        return get_total1

    @allure.step("User Management页面，导入前获取列表筛选的User ID，如果能筛选到1条记录，先删除已存在的用户")
    def delete_repetitive_user(self):
        sql1 = SQL('DCR', 'test')
        try:
            varsql1 = "select USER_CODE from t_user where USER_CODE ='smarttest102'"
            varsql2 = "select EMP_CODE from t_employee where EMP_CODE ='smarttest102'"
            result_user = sql1.query_db(varsql1)
            result_emp = sql1.query_db(varsql2)
            user_code = result_user[0].get("USER_CODE")
            emp_code = result_emp[0].get("EMP_CODE")
            if user_code == 'smarttest102':
                """ 在数据库表中，删除导入的用户 """
                sql1.delete_db(
                    "delete from t_user where USER_CODE ='smarttest102'")
            elif emp_code == 'smarttest102':
                sql1.delete_db(
                    "delete from t_employee where EMP_CODE ='smarttest102'")
            else:
                logging.info("获取User Management页面user_code与emp_code字段内容：{}".format(user_code, emp_code))
        except Exception as e:
            logging.info("如有异常打印日志{}".format(e))


    @allure.step("导入用户模板-上传正确的文件")
    def upload_true_file(self, file1):
        path1 = os.path.join(BASE_DIR, 'project', 'DCR', 'data', file1)
        logging.info("打印上传的用户模板文件path：{}".format(path1))
        self.click_import_upload_save(path1)

    @allure.step("导入记录页面，根据Import Date条件筛选导入开始日期")
    def import_record_import_date_query(self, start_date):
        self.is_click(user['Query Import Start Date'])
        self.input_text(user['Query Import Start Date'], start_date)

    @allure.step("循环点击查询，直到获取到导入记录状态为Upload Successfully")
    def click_import_status_search(self):
        import_status = self.import_record_status(user['Import Search'], user['Get Import Record Status'])
        return import_status

    """导入记录页面，获取列表字段断言是否导入成功"""
    @allure.step("Import Record页面，获取File Name字段文本")
    def get_import_file_name(self):
        get_file_name = self.element_text(user['Get Import Record File Name'])
        return get_file_name

    @allure.step("Import Record页面，获取Status字段文本")
    def get_import_status(self):
        get_status = self.element_text(user['Get Import Record Status'])
        return get_status

    @allure.step("Import Record页面，获取Total字段文本")
    def get_import_total(self):
        get_total = self.element_text(user['Get Import Record Total'])
        return get_total

    @allure.step("Import Record页面，获取Total字段文本")
    def get_import_success(self):
        get_success = self.element_text(user['Get Import Record Success'])
        return get_success

    @allure.step("Import Record页面，获取Failed字段文本")
    def get_import_failed(self):
        get_failed = self.element_text(user['Get Import Record Failed'])
        return get_failed

    @allure.step("Import Record页面，获取 Fail Data字段文本")
    def get_import_fail_data(self):
        self.scroll_into_view(user['Get Import Fail Data'])
        get_fail_data = self.element_text(user['Get Import Fail Data'])
        return get_fail_data

    @allure.step("Import Record页面，获取 Import Date字段文本")
    def get_import_import_date(self):
        self.scroll_into_view(user['Get Import Import Date'])
        get_import_date = self.element_text(user['Get Import Import Date'])
        get_import_date1 = get_import_date[0:10]
        return get_import_date1

    @allure.step("User Management页面页面，获取列表Brand字段内容")
    def get_list_brand(self):
        get_brand = self.element_text(user['Get list Brand'])
        return get_brand

    @allure.step("User Management页面页面，获取列表 Country字段内容")
    def get_list_country(self):
        sleep(1.5)
        self.scroll_into_view(user['Get list Country'])
        get_country = self.element_text(user['Get list Country'])
        return get_country

    @allure.step("User Management页面页面，获取列表 Position字段内容")
    def get_list_position(self):
        sleep(1.5)
        self.scroll_into_view(user['Get list Position'])
        get_position = self.element_text(user['Get list Position'])
        return get_position

    @allure.step("User Management页面页面，获取列表 Staff Status字段内容")
    def get_list_staff_status(self):
        get_staff_status = self.element_text(user['Get list Staff Status'])
        return get_staff_status

    """导出用户"""
    @allure.step("User Management页面，点击Export 导出用户记录")
    def click_export(self):
        self.is_click(user['Export'])
        sleep(1.5)

    @allure.step("User Management页面，导出操作后，点击右上角下载图标,点击右上角more...")
    def click_download_more(self):
        self.mouse_hover_click(user['Download Icon'])
        Base.presence_sleep_dcr(self, user['More'])
        self.is_click(user['More'])
        sleep(3)

    @allure.step("输入Task Name筛选该任务的导出记录")
    def input_task_name(self, task_name):
        self.is_click(user['Input Task Name'])
        self.input_text(user['Input Task Name'], task_name)
        sleep(0.5)
        self.presence_sleep_dcr(user['Input Task Name value'], task_name)
        self.is_click(user['Input Task Name value'], task_name)

    @allure.step("输入Create Date开始日期筛选当天日期的导出记录")
    def export_record_create_start_date(self, start_date):
        self.is_click(user['导出记录筛选创建日期'])
        self.input_text(user['导出记录筛选创建日期'], start_date)
        self.is_click(user['点击筛选条件的标签'], 'Create Date')

    @allure.step("循环点击查询，直到获取到下载状态为COMPLETE")
    def click_export_search(self):
        download_status = self.export_download_status(user['Export Record Search'], user['获取下载状态文本'])
        return download_status

    @allure.step("导出记录页面，获取列表 Task Name文本")
    def get_task_name_text(self):
        task_name = self.element_text(user['获取任务名称文本'])
        return task_name

    @allure.step("导出记录页面，获取列表 Task Name文本")
    def get_file_size_text(self):
        file_size = self.element_text(user['获取文件大小文本'])
        file_size1 = file_size[0:1]
        return file_size1

    @allure.step("导出记录页面，获取列表 User ID文本")
    def get_task_user_id_text(self):
        user_id = self.element_text(user['获取用户ID文本'])
        return user_id

    @allure.step("导出记录页面，获取列表 Create Date文本")
    def get_create_date_text(self):
        self.scroll_into_view(user['获取创建日期文本'])
        create_date = self.element_text(user['获取创建日期文本'])
        create_date1 = create_date[0:10]
        return create_date1

    @allure.step("导出记录页面，获取列表Complete Date文本")
    def get_complete_date_text(self):
        complete_date = self.element_text(user['获取完成日期文本'])
        complete_date1 = complete_date[0:10]
        return complete_date1

    @allure.step("导出记录页面，获取列表 Operation文本")
    def get_operation_text(self):
        operation = self.element_text(user['获取操作按钮文本'])
        return operation

    @allure.step("导出记录页面，获取列表导出时间文本")
    def get_export_time_text(self):
        export_time = self.element_text(user['获取导出时间'])
        export_time1 = export_time[0:1]
        return export_time1

    @allure.step("断言分页总数是否存在数据")
    def assert_total(self, total):
        if int(total) > 0:
            logging.info("筛选考勤记录列表，分页总条数大于0，能查询到考勤记录数Total:{}".format(total))
        else:
            logging.info("筛选考勤记录列表，分页总条数为0，未查询到考勤记录数Total:{}:".format(total))

    @allure.step("断言分页总数是否存在数据")
    def assert_total2(self, total):
        if int(total) > 1000:
            logging.info("查看考勤记录列表，分页总条数大于1000，能查询到考勤记录Total：{}".format(total))
        else:
            logging.info("查看考勤记录列表，分页总条数为1000，未查询到考勤记录Total：{}".format(total))

    @allure.step("断言文件或导出时间是否有数据")
    def assert_file_time_size(self, file_size, export_time):
        if int(file_size) > 0:
            logging.info("Attendance Records导出成功，File Size导出文件大于M:{}".format(file_size))
        else:
            logging.info("Attendance Records导出失败，File Size导出文件小于M:{}".format(file_size))

        if int(export_time) > 0:
            logging.info("Attendance Records导出成功，Export Time(s)导出时间大于0s:{}".format(export_time))
        else:
            logging.info("Attendance Records导出失败，Export Time(s)导出时间小于0s:{}".format(export_time))

    @allure.step("关闭导出记录菜单")
    def click_close_export_record(self):
        """关闭导出记录菜单"""
        self.is_click(user['关闭导出记录菜单'])

    @allure.step("点击Unfold 展开筛选项")
    def click_unfold(self):
        self.is_click(user['Unfold'])
        logging.info('点击Unfold 展开筛选项')

    @allure.step("user management页面，输入查询条件")
    def input_search(self, header, content):
        """
        @header： 输入框名称
        @content： 输入内容
        """
        user_list = ['User ID']
        country_list = ['Sales Region', 'Country/City']
        fuzzySelect_list = ['Belong To Customer', 'Superior']
        exactSelect_list = ['Staff Status', 'Have Superior or Not', 'Have Shop or Not', 'Staff Type']
        inputSelect_list = ['Brand', 'Position', 'Role']
        self.element_exist(user['Loading'])
        logging.info(f'输入查询条件： {header} ，内容： {content}')
        if header in user_list:
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框3'], content, header)
        elif header == 'User Name':
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框'], content, header)
            self.is_click_tbm(user['输入结果模糊选择'], content)
        elif header in exactSelect_list:
            self.is_click_tbm(user['输入框'], header)
            self.is_click_tbm(user['输入结果精确选择'], content)
        elif header in fuzzySelect_list:
            if content != '':
                self.is_click_tbm(user['输入框'], header)
                self.input_text(user['输入框2'], content, header)
                self.is_click_tbm(user['输入结果模糊选择'], content)
        elif header in country_list:
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框'], content, header)
            self.is_click_tbm(user['地区选择框'], content)
        elif header in inputSelect_list:
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框4'], content, header)
            self.is_click_tbm(user['输入结果精确选择'], content)
        else:
            logging.error('请输入正确的查询条件')
            raise ValueError('请输入正确的查询条件')

    @allure.step("创建页面输入Job_Information")
    def input_Job_Information(self, header, content):
        """
        @header： 输入框名称
        @content： 输入内容
        """
        user_list = ['User ID', 'User Name']
        country_list = ['Sales Region', 'Country/City']
        FuzzySelect_list = ['Belong To Customer', 'Superior']
        ExactSelect_list = ['Position']
        logging.info(f'输入查询条件： {header} ，内容： {content}')
        if header in user_list:
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框'], content, header)
        elif header in country_list:
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框'], content, header)
            self.is_click_tbm(user['地区选择框'], content)
        elif header in FuzzySelect_list:
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框'], content, header)
            self.is_click_tbm(user['输入结果模糊选择'], content)
        elif header in ExactSelect_list:
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框'], content, header)
            self.is_click_tbm(user['输入结果精确选择'], content)
        elif header == 'Brand':
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框4'], content, header)
            self.is_click_tbm(user['输入结果精确选择'], content)
        elif header == 'Staff Type':
            self.is_click_tbm(user['输入框'], header)
            self.is_click_tbm(user['输入结果精确选择'], content)
        elif header == 'Shop':
            self.is_click_tbm(user['输入框'], header)
            self.input_text(user['输入框2'], content, header)
            self.is_click_tbm(user['输入结果模糊选择'], content)
        else:
            logging.error('请输入正确的查询条件')
            raise ValueError('请输入正确的查询条件')
        self.is_click_tbm(user['JobInformation标题'])

    @allure.step("创建页面输入Personal_Information")
    def input_Personal_Information(self, header, content):
        """
        @header： 输入框名称
        @content： 输入内容
        """
        user_list = ['User ID', 'User Name']
        country_list = ['Sales Region', 'Country/City']
        FuzzySelect_list = ['Belong To Customer', 'Superior']
        ExactSelect_list = ['Position']
        logging.info(f'输入查询条件： {header} ，内容： {content}')
        if header == 'Gender':
            self.is_click_tbm(user['输入框'], header)
            self.is_click_tbm(user['输入结果精确选择'], content)

    @allure.step("判断空值")
    def assert_None(self, result):
        try:
            assert result == '', logging.warning("断言失败: 该值不为None | x:{}".format(result))
            logging.info("断言成功: 该值为None | x:{}".format(result))
        except Exception as e:
            logging.error(e)
            raise

    @allure.step("断言：页面查询结果")
    def assert_User_Exist(self, header, content):
        logging.info('开始断言：页面查询结果')
        DomAssert(self.driver).assert_search_contains_result(user['menu表格字段'], user['表格内容'], header, content, sc_element=user['滚动条'], h_element=user['表头文本'])

    @allure.step("断言：页面查询结果")
    def assert_search_result(self, header, content):
        logging.info(f'开始断言：页面查询：{header} 结果 ：{content}')
        if header == 'Superior' or header == 'Belong To Customer':
            self.assert_User_Exist(f'{header} ID', content)
        elif header == 'Have Superior or Not' or header == 'Have Shop or Not':
            column = self.get_table_info(user['menu表格字段'], f"{header.split(' ')[1]} ID", sc_element=user['滚动条'], h_element=user['表头文本'])
            contents = self.get_row_info(user['表格内容'], column, user['滚动条'])
            if content == 'Yes':
                for i in contents:
                    ValueAssert.value_assert_IsNoneNot(i)
            else:
                for i in contents:
                    self.assert_None(i)
        elif header == 'Sales Region':
            self.assert_User_Exist(f'{header}5', content)
        elif header == 'Country/City':
            self.assert_User_Exist(f'City', content)
        elif header == 'Staff Type':
            column = self.get_table_info(user['menu表格字段'], 'Belong To Customer ID', sc_element=user['滚动条'], h_element=user['表头文本'])
            contents = self.get_row_info(user['表格内容'], column, user['滚动条'])
            if content == 'Dealer Staff':
                for i in contents:
                    ValueAssert.value_assert_IsNoneNot(i)
            elif content == 'Transsion Staff':
                for i in contents:
                    self.assert_None(i)
        else:
            self.assert_User_Exist(header, content)

    @allure.step("点击复选框")
    def click_checkbox(self, UID, header='User ID'):
        """
        @UID： UID 默认传入userid，username或其他唯一内容也可以传
        """
        column = self.get_table_info(user['表格字段'], header, h_element=user['表头文本'])
        self.is_click_tbm(user['指定复选框'], column, UID)
        logging.info(f'点击 {UID} 复选框')

    @allure.step("点击指定编辑")
    def click_Edit(self, UID, header='User ID'):
        """
        @UID： UID 默认传入userid，如传入的是username或其他，则需要修改header
        @header： 默认是'User ID'，表格表头，如：传name就用User Name
        """
        column = self.get_table_info(user['表格字段'], header, h_element=user['表头文本'])
        self.is_click_tbm(user['指定编辑'], column, UID)
        logging.info('点击指定编辑')
        self.element_exist(user['Loading'])

    @allure.step("断言：输入框是否可以编辑")
    def assert_input_edit(self, header, result=False):
        """
        @header： 输入框名称
        @result： 默认False 表示不可编辑；True表示可编辑
        """
        logging.info(f'开始断言：{header} 输入框是否可以编辑')
        # acresult = self.get_element_attribute(user['输入框'], 'disabled', header)
        acresult = self.find_element(user['输入框'], header).is_enabled()
        try:
            assert acresult == result
            logging.info(f'断言成功，{header} 输入框编辑状态是 {acresult} ，与期望结果 {result} 一致')
        except:
            logging.info(f'断言失败，{header} 输入框编辑状态是 {acresult} ，与期望结果 {result} 不一致')
            raise

    @allure.step("查找菜单")
    def click_menu(self, *content):
        self.refresh()
        self.is_click_tbm(user['菜单栏'])
        self.refresh()
        for i in range(len(content)):
            self.is_click_tbm(user['菜单'], content[i])
            logging.info('点击菜单：{}'.format(content[i]))
        self.refresh()
        self.element_exist(user['Loading'])

    @allure.step("点击Upload按钮")
    def click_upload(self):
        self.is_click(user['Upload'])
        logging.info('点击upload按钮')
        # k = PyKeyboard()
        # k.tap_key(k.escape_key)

    @allure.step("点击EditUpload按钮")
    def click_EditUpload(self):
        self.is_click(user['EditUpload'])
        logging.info('点击upload按钮')

    @allure.step("点击AddUpload按钮")
    def click_AddUpload(self):
        self.is_click(user['AddUpload'])
        logging.info('点击upload按钮')

    @allure.step("点击Import按钮")
    def click_AddImport(self):
        self.is_click(user['Import'])
        logging.info('点击Import按钮')
        self.click_AddUpload()

    @allure.step("点击Import按钮")
    def click_EditImport(self):
        self.is_click(user['Import'])
        logging.info('点击Import按钮')
        self.click_EditUpload()

    @allure.step("点击Save按钮")
    def click_save(self):
        self.is_click(user['Save'])
        logging.info('点击Save按钮')

    @allure.step("点击Confirm按钮")
    def click_confirm(self):
        self.is_click(user['Confirm'])
        logging.info('点击Confirm按钮')
        sleep(2)
        self.refresh()

    @allure.step("断言：导入成功状态")
    def assert_import_success(self):
        logging.info("开始断言：导入成功状态")
        DomAssert(self.driver).assert_control(user['导入成功状态'])

    @allure.step("导入文件")
    def import_file(self, name):
        """
        @name： 传入存放在data文件夹里的文件名
        """
        file_path = os.path.join(BASE_DIR, 'project', 'DCR', 'data', name)
        logging.info("文件地址：{}".format(file_path))
        self.upload_file(user['导入'], file_path)
        logging.info("导入文件：{}".format(file_path))

    @allure.step("导入文件")
    def EditImport_file(self, name):
        """
        @name： 传入存放在data文件夹里的文件名
        """
        file_path = os.path.join(BASE_DIR, 'project', 'DCR', 'data', name)
        logging.info("文件地址：{}".format(file_path))
        self.upload_file(user['导入'], file_path)
        logging.info("导入文件：{}".format(file_path))

    @allure.step("获得Record指定内容")
    def get_Record_info(self, menu, name, header):
        """
        :param menu: 菜单名
        :param name: 输入文件名
        :param header: 需要获取的指定字段
        """
        for i in range(20):
            ac_menu = self.element_text(user['当前菜单'])
            if ac_menu == menu:
                column = self.get_table_info(user['表格字段'], header)
                content = self.element_text(user['表格指定列内容'], name, column)
                logging.info(f'获得 {menu} 页面 {name} 文件 {header} 字段内容 {content}')
                return content
            self.click_menu('Basic Data Management', menu)
            column = self.get_table_info(user['表格字段'], header)
            content = self.element_text(user['表格指定列内容'], name, column)
            logging.info(f'获得 {menu} 页面 {name} 文件 {header} 字段内容 {content}')
            return content

    @allure.step("检查导入导出记录状态是否为Upload Successfully")
    def check_Record_result(self, menu, name):
        """
        :param menu: 菜单
        :param name: 输入文件名
        """
        for i in range(20):
            RecordResult = self.get_Record_info(menu, name, 'Status')
            if RecordResult == 'Execute':
                logging.info(f'Status结果为{RecordResult}，刷新更新结果状态')
                self.click_search()
                sleep(1)
            elif RecordResult == 'Upload Successfully':
                logging.info(f'Status结果为{RecordResult}')
                return
            else:
                logging.error('上传结果异常，请检查')
                raise ValueError('上传结果异常，请检查')

    @allure.step("断言：导入导出Record结果")
    def assert_Record_result(self, menu, name, header, result=None):
        """
        :param menu: 菜单
        :param name: 输入文件名
        :param header: 需要获取的指定字段
        :param result: 需要断言的值 比如状态，数量，时间
        """
        logging.info('开始断言：导入导出Record结果')
        self.check_Record_result(menu, name)
        ac_result = self.get_Record_info(menu, name, header)
        if header == 'File Size':
            ValueAssert.value_assert_IsNot(ac_result, '0B')
            logging.info(f'断言成功：{menu} 页面 {name}文件的FileSize实际结果{ac_result}，大于0B')
        else:
            ValueAssert.value_assert_In(result, ac_result)
            logging.info(f'断言成功：{menu} 页面{name}文件{header} 字段实际结果{ac_result} 与期望结果{result}一致')

    @allure.step("获得Edit页面Information")
    def get_Edit_Information(self, header):
        """
        :param header: 需要获取的指定字段
        """
        Information = self.element_input_text(user['输入框'], header)
        logging.info(f'获得Edit页面{header} 字段 Information：{Information}')
        return Information

    @allure.step("断言：页面Information一致")
    def assert_user_Information(self, header, content):
        """
        :param header: 需要断言的字段名
        :param content: 需要断言的内容
        """
        logging.info(f'开始断言：Edit页面Information一致')
        ValueAssert.value_assert_equal(content, self.get_Edit_Information(header))

    @allure.step("导入员工文件编辑")
    def Edit_User_file(self, name, UID, header, content):
        """
        :param name: 输入文件名
        :param UID: Excel表格内容的user id
        :param header: 表头字段
        :param content: 需要修改的内容
        """
        logging.info(f'开始编辑 {name} Excel文件')
        file_path = os.path.join(BASE_DIR, 'project', 'DCR', 'data', name)
        logging.info("文件地址：{}".format(file_path))
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        id_cells = sheet['B']
        for i in id_cells:
            if str(i.value) == UID:
                logging.info(f'获取excle表格指定UID：坐标：{i.coordinate} : {i.value}')
                header_cells = sheet['1']
                for j in header_cells:
                    if j.value == header:
                        sheet.cell(row=i.row, column=j.column).value = content
                        logging.info(f'修改行：{i.row}，列：{j.column} 内容：{content}')
        workbook.save(filename=file_path)

    @allure.step("点击取消")
    def click_Cancel(self):
        self.is_click_tbm(user['Cancel'])
        logging.info("点击取消")

    @allure.step("点击Add")
    def click_Add(self):
        self.is_click_tbm(user['Add1'])
        logging.info("点击Add")

    @allure.step("点击功能按钮")
    def click_function_button(self, function, confirm='Yes'):
        """
        @function： 需要点击的功能按钮，具体如下：
        Add, Import, Export, More Option,
        Enable, Reset Password, Quit
        """
        MoreOptionList = ['Enable', 'Reset Password', 'Quit']
        if function in MoreOptionList:
            self.is_click_tbm(user['功能按钮'], 'More Option')
            self.is_click_tbm(user['功能按钮2'], function)
            if function == 'Quit':
                self.is_click_tbm(user['UserDisableSettingYes'])
            elif function == 'Enable':
                self.is_click_tbm(user['EnableEmployeesYes'])
            elif function == 'Reset Password':
                if confirm == 'Yes':
                    self.is_click_tbm(user['ResetPasswordYes'])
        else:
            self.is_click_tbm(user['功能按钮'], function)
        logging.info(f'点击功能按钮： {function}')

    @allure.step("复职用户 组合方法")
    def enable_user_Method(self, uid):
        logging.info('开始使用组合方法: 复职用户')
        self.click_unfold()
        self.input_search('User ID', uid)
        self.click_search()
        total_text = self.element_text(user['Total'])
        total = total_text[total_text.index(' ')+1:]
        logging.info(f'查询结果合计数{total_text}')
        if total != '0':
            self.click_checkbox(uid)
            self.click_function_button('Quit')
            DomAssert(self.driver).assert_att('Disabled Successfully')
        self.input_search('Staff Status', 'Off Service')
        self.click_search()
        self.click_checkbox(uid)
        self.click_function_button('Enable')
        DomAssert(self.driver).assert_att('Set Up Successfully')
        self.refresh()
        sleep(10)

    @allure.step("停职用户 组合方法")
    def disable_user_Method(self, uid):
        logging.info('开始使用组合方法: 停职用户')
        self.input_search('User ID', uid)
        self.click_search()
        total_text = self.element_text(user['Total'])
        total = total_text[total_text.index(' ')+1:]
        logging.info(f'查询结果合计数{total_text}')
        if total != '0':
            self.click_checkbox(uid)
            self.click_function_button('Quit')
            DomAssert(self.driver).assert_att('Disabled Successfully')
            self.refresh()

    @allure.step("数据库删除指定用户")
    def SQL_delete_user(self, uid):
        logging.info(f'数据库删除 {uid} 用户')
        a = SQL('DCR', 'test')
        a.change_db(
            f"delete from t_user where USER_CODE = {uid}"
        )
        a.change_db(
            f"delete from t_employee where EMP_CODE = {uid}"
        )

    @allure.step("组合查询 组合方法")
    def random_Query_Method(self, kwargs):
        list_query = []
        num = random.randint(3, 8)
        for i in kwargs:
            list_query.append(i)
        logging.info(f'输入框：{list_query}')
        list_random = random_list(list_query, num)
        logging.info(f'随机组合：输入框：{list_random}')
        for i in list_random:
            logging.info(f'随机组合：{i} 输入框输入内容：{kwargs[i]}')
            self.input_search(i, kwargs[i])
        self.click_search()
        for i in list_random:
            self.assert_search_result(i, kwargs[i])

    @allure.step("修改密码")
    def change_pwd(self, password):
        self.input_text(user['新密码'], password)
        self.input_text(user['确认密码'], password)
        self.is_click_tbm(user['修改密码保存'])
        DomAssert(self.driver).assert_att('Save successfully!')
        self.is_click_tbm(user['修改密码OK'])



if __name__ == '__main__':
    pass